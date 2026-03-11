"""Data export formats for FreeClimber.

Supports: slopes CSV (backward compatible), tidy CSV (R-ready),
Excel workbook, GraphPad Prism format, per-fly tracks.
"""

import logging
import os

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)


def export_slopes_csv(slopes_df: pd.DataFrame, path: str):
    """Export slopes in original FreeClimber format (backward compatible)."""
    slopes_df.to_csv(path, index=False)
    logger.info(f'Saved slopes CSV: {path}')


def export_tidy_csv(slopes_df: pd.DataFrame, path: str, experiment: str = ''):
    """Export in R-ready long/tidy format (one observation per row).

    Columns: experiment, vial_ID, metric_name, metric_value, plus any
    experimental detail columns (genotype, sex, day, rep, etc.)
    """
    # Identify experimental detail columns vs metric columns
    metric_cols = ['slope', 'intercept', 'r_value', 'p_value', 'std_err',
                   'first_frame', 'last_frame']
    detail_cols = [c for c in slopes_df.columns if c not in metric_cols and c != 'vial_ID']

    rows = []
    for _, row in slopes_df.iterrows():
        base = {'experiment': experiment, 'vial_ID': row.get('vial_ID', '')}
        for dc in detail_cols:
            if dc in row.index:
                base[dc] = row[dc]

        for mc in metric_cols:
            if mc in row.index:
                r = base.copy()
                r['metric_name'] = mc
                r['metric_value'] = row[mc]
                rows.append(r)

    tidy_df = pd.DataFrame(rows)
    tidy_df.to_csv(path, index=False)
    logger.info(f'Saved tidy CSV: {path}')


def export_excel(slopes_df: pd.DataFrame, path: str,
                 stats_df: pd.DataFrame = None,
                 per_fly_df: pd.DataFrame = None,
                 raw_df: pd.DataFrame = None,
                 params: dict = None):
    """Export professional Excel workbook with multiple sheets.

    Sheet 1: Summary (slopes with formatting)
    Sheet 2: Statistics (if provided)
    Sheet 3: Per-Fly Metrics (if individual tracking)
    Sheet 4: Raw Particles (if provided)
    Sheet 5: Parameters
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed; saving as plain CSV instead")
        slopes_df.to_csv(path.replace('.xlsx', '.csv'), index=False)
        return

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # Sheet 1: Summary
        slopes_df.to_excel(writer, sheet_name='Summary', index=False)

        # Sheet 2: Statistics
        if stats_df is not None:
            stats_df.to_excel(writer, sheet_name='Statistics', index=False)

        # Sheet 3: Per-Fly Metrics
        if per_fly_df is not None and not per_fly_df.empty:
            per_fly_df.to_excel(writer, sheet_name='Per-Fly Metrics', index=False)

        # Sheet 4: Raw Particles
        if raw_df is not None:
            # Limit to first 50000 rows to avoid huge files
            raw_df.head(50000).to_excel(writer, sheet_name='Raw Particles', index=False)

        # Sheet 5: Parameters
        if params:
            params_df = pd.DataFrame([
                {'Parameter': k, 'Value': str(v)} for k, v in params.items()
            ])
            params_df.to_excel(writer, sheet_name='Parameters', index=False)

    # Apply formatting
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb['Summary']

        # Bold headers
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        # Auto-size columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

        wb.save(path)
    except Exception as e:
        logger.warning(f'Could not apply Excel formatting: {e}')

    logger.info(f'Saved Excel workbook: {path}')


def export_prism_csv(slopes_df: pd.DataFrame, path: str,
                     group_col: str = None, value_col: str = 'slope'):
    """Export in GraphPad Prism-compatible grouped column format.

    Genotype as column headers, values as rows.
    """
    if group_col is None:
        # Try to detect group column
        for candidate in ['geno', 'genotype', 'group', 'condition']:
            if candidate in slopes_df.columns:
                group_col = candidate
                break

    if group_col is None or group_col not in slopes_df.columns:
        # Fall back to vial_ID
        group_col = 'vial_ID'

    # Pivot: each group becomes a column
    groups = slopes_df.groupby(group_col)[value_col].apply(list).to_dict()
    max_len = max(len(v) for v in groups.values())

    prism_df = pd.DataFrame({
        name: values + [np.nan] * (max_len - len(values))
        for name, values in groups.items()
    })

    prism_df.to_csv(path, index=False)
    logger.info(f'Saved Prism CSV: {path}')


def export_per_fly_tracks(df: pd.DataFrame, path: str):
    """Export individual fly tracks: fly_id, frame, x, y, speed, vial."""
    if 'particle' not in df.columns:
        logger.warning("No individual tracking data; skipping per-fly export")
        return

    cols = ['particle', 'frame', 'x', 'y', 'vial']
    available = [c for c in cols if c in df.columns]
    export_df = df[available].copy()
    export_df = export_df.rename(columns={'particle': 'fly_id'})
    export_df.to_csv(path, index=False)
    logger.info(f'Saved per-fly tracks: {path}')
